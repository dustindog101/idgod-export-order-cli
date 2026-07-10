from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .models import (
    EXPORT_ONLY_FIELDS,
    FIELD_ALIASES,
    ExportBundle,
    ExportMeta,
    OrderBatch,
    Person,
    ShippingInfo,
)
from .states import expand_state_name, variant_from_product_id

LOCAL_DELIVERY = "local delivery"

JSON_ID_FIELD_MAP: dict[str, str] = {
    "productid": "product_id",
    "state": "state",
    "firstname": "first_name",
    "middlename": "middle_name",
    "lastname": "last_name",
    "dob": "dob",
    "issuedate": "issue_date",
    "streetaddress": "street",
    "city": "city",
    "zipcode": "zip",
    "zipplus4": "zip4",
    "sex": "sex",
    "height": "height",
    "weight": "weight",
    "eyecolor": "eye_color",
    "haircolor": "hair_color",
    "photourl": "photo",
    "signatureurl": "signature",
}


def _norm_key(key: str) -> str:
    return re.sub(r"\s+", " ", key.strip().lower())


def _clean(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_local_delivery(text: str) -> bool:
    return _clean(text).lower() == LOCAL_DELIVERY


def _mapped_person_fields(row: dict[str, Any]) -> tuple[dict[str, str], str]:
    mapped: dict[str, str] = {}
    export_order_id = ""
    for raw_key, raw_val in row.items():
        key = _norm_key(str(raw_key))
        if key in EXPORT_ONLY_FIELDS:
            if key == "order id" and raw_val:
                export_order_id = _clean(raw_val)
            continue
        field = FIELD_ALIASES.get(key)
        if field and raw_val is not None and _clean(raw_val):
            mapped[field] = _clean(raw_val)
    return mapped, export_order_id


def _person_from_mapped(
    mapped: dict[str, str],
    *,
    source_row: int | None = None,
    export_order_id: str = "",
    shipping_raw: str = "",
    product_id: str = "",
) -> Person:
    required = ("first_name", "last_name", "state", "dob", "city", "zip")
    missing = [r for r in required if not mapped.get(r)]
    if missing:
        raise ValueError(f"Row {source_row or '?'} missing required fields: {', '.join(missing)}")

    pid = product_id or mapped.get("product_id", "")
    state_variant = mapped.get("state_variant", "") or variant_from_product_id(pid)
    shipping = _clean(shipping_raw)

    return Person(
        first_name=mapped["first_name"],
        last_name=mapped["last_name"],
        state=expand_state_name(mapped["state"]),
        dob=mapped["dob"],
        city=mapped["city"],
        zip=mapped["zip"],
        middle_name=mapped.get("middle_name", ""),
        issue_date=mapped.get("issue_date", ""),
        street=mapped.get("street", ""),
        zip4=mapped.get("zip4", ""),
        sex=mapped.get("sex", ""),
        height=mapped.get("height", ""),
        weight=mapped.get("weight", ""),
        eye_color=mapped.get("eye_color", ""),
        hair_color=mapped.get("hair_color", ""),
        photo=mapped.get("photo", ""),
        signature=mapped.get("signature", ""),
        state_variant=state_variant,
        product_id=pid,
        email=mapped.get("email", ""),
        shipping_raw=shipping,
        local_delivery=_is_local_delivery(shipping),
        source_row=source_row,
        export_order_id=export_order_id,
    )


def _row_to_person(row: dict[str, Any], source_row: int | None = None) -> Person:
    mapped, export_order_id = _mapped_person_fields(row)
    shipping_raw = ""
    for raw_key, raw_val in row.items():
        key = _norm_key(str(raw_key))
        if key in ("shipping", "shipping address") and raw_val is not None and _clean(raw_val):
            shipping_raw = _clean(raw_val)
            break
    product_id = ""
    for raw_key, raw_val in row.items():
        if _norm_key(str(raw_key)) == "product id" and raw_val is not None:
            product_id = _clean(raw_val)
            break
    return _person_from_mapped(
        mapped,
        source_row=source_row,
        export_order_id=export_order_id,
        shipping_raw=shipping_raw,
        product_id=product_id,
    )


def _json_id_to_mapped(id_row: dict[str, Any]) -> dict[str, str]:
    mapped: dict[str, str] = {}
    for raw_key, raw_val in id_row.items():
        norm = re.sub(r"[^a-z0-9]+", "", str(raw_key).lower())
        field = JSON_ID_FIELD_MAP.get(norm)
        if field and raw_val is not None and _clean(raw_val):
            mapped[field] = _clean(raw_val)
    return mapped


def _parse_export_json_data(data: dict[str, Any]) -> ExportBundle:
    meta = ExportMeta(
        exported_at=_clean(data.get("exportedAt")),
        export_note=_clean(data.get("exportNote") or ""),
        shipping_override=data.get("shippingOverride"),
        order_count=int(data.get("orderCount") or 0),
        id_row_count=int(data.get("idRowCount") or 0),
    )
    override = _clean(meta.shipping_override or "")
    batches: list[OrderBatch] = []
    source_row = 2

    for order in data.get("orders") or []:
        order_id = _clean(order.get("orderId"))
        shipping_raw = override or _clean(order.get("shippingAddress"))
        local_delivery = _is_local_delivery(shipping_raw)
        people: list[Person] = []

        for id_row in order.get("ids") or []:
            mapped = _json_id_to_mapped(id_row)
            product_id = _clean(id_row.get("productId"))
            people.append(
                _person_from_mapped(
                    mapped,
                    source_row=source_row,
                    export_order_id=order_id,
                    shipping_raw=shipping_raw,
                    product_id=product_id,
                )
            )
            source_row += 1

        batches.append(
            OrderBatch(
                order_id=order_id,
                people=people,
                shipping_raw=shipping_raw,
                local_delivery=local_delivery,
                status=_clean(order.get("status")),
                order_note=_clean(order.get("orderNote")),
                export_note=_clean(order.get("exportNote")),
                tracking_number=_clean(order.get("trackingNumber")),
            )
        )

    if not meta.order_count:
        meta.order_count = len(batches)
    if not meta.id_row_count:
        meta.id_row_count = sum(b.id_count for b in batches)
    return ExportBundle(meta=meta, batches=batches)


def _xlsx_sheet(wb: Any) -> Any:
    if "Orders" in wb.sheetnames:
        return wb["Orders"]
    return wb[wb.sheetnames[0]]


def _parse_export_xlsx(path: Path) -> ExportBundle:
    try:
        import openpyxl
    except ImportError as e:
        raise RuntimeError("openpyxl required: pip install openpyxl") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = _xlsx_sheet(wb)
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ExportBundle(meta=ExportMeta(), batches=[])

    headers = [str(h or "").strip() for h in rows[0]]
    batches_by_order: dict[str, OrderBatch] = {}
    order_sequence: list[str] = []
    source_row = 2

    for row in rows[1:]:
        if not row or all(v is None or _clean(v) == "" for v in row):
            source_row += 1
            continue

        data = {headers[j]: row[j] if j < len(row) else "" for j in range(len(headers))}
        order_id = _clean(data.get("Order ID"))
        id_index = data.get("ID #")
        has_id = id_index not in (None, "", 0)

        if not has_id:
            if order_id and order_id not in batches_by_order:
                shipping_raw = _clean(data.get("Shipping Address"))
                batches_by_order[order_id] = OrderBatch(
                    order_id=order_id,
                    people=[],
                    shipping_raw=shipping_raw,
                    local_delivery=_is_local_delivery(shipping_raw),
                    status=_clean(data.get("Status")),
                    order_note=_clean(data.get("Order Note")),
                    export_note=_clean(data.get("Export Note")),
                    tracking_number=_clean(data.get("Tracking #")),
                )
                order_sequence.append(order_id)
            source_row += 1
            continue

        person = _row_to_person(data, source_row=source_row)
        key = order_id or f"row-{source_row}"
        if key not in batches_by_order:
            batches_by_order[key] = OrderBatch(
                order_id=order_id,
                people=[],
                shipping_raw=person.shipping_raw,
                local_delivery=person.local_delivery,
                status=_clean(data.get("Status")),
                order_note=_clean(data.get("Order Note")),
                export_note=_clean(data.get("Export Note")),
                tracking_number=_clean(data.get("Tracking #")),
            )
            order_sequence.append(key)
        batches_by_order[key].people.append(person)
        source_row += 1

    batches = [batches_by_order[k] for k in order_sequence if k in batches_by_order]
    return ExportBundle(
        meta=ExportMeta(
            order_count=len(batches),
            id_row_count=sum(b.id_count for b in batches),
        ),
        batches=batches,
    )


def parse_export_file(path: Path) -> ExportBundle:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return _parse_export_xlsx(path)
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict) and "orders" in data:
            return _parse_export_json_data(data)
        people = parse_json(path)
        if not people:
            return ExportBundle(meta=ExportMeta(), batches=[])
        return ExportBundle(
            meta=ExportMeta(order_count=1, id_row_count=len(people)),
            batches=[OrderBatch(order_id=people[0].export_order_id, people=people)],
        )
    if suffix == ".csv":
        people = parse_csv(path)
        shipping_raw = ""
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                shipping_raw = _shipping_from_row(row)
                if shipping_raw:
                    break
        for person in people:
            if not person.shipping_raw:
                person.shipping_raw = shipping_raw
                person.local_delivery = _is_local_delivery(shipping_raw)
        return ExportBundle(
            meta=ExportMeta(order_count=1, id_row_count=len(people)),
            batches=[
                OrderBatch(
                    order_id=people[0].export_order_id if people else "",
                    people=people,
                    shipping_raw=shipping_raw,
                    local_delivery=_is_local_delivery(shipping_raw),
                )
            ],
        )
    raise ValueError(f"Unsupported file type: {suffix} (use .csv, .xlsx, or .json)")


def parse_csv(path: Path) -> list[Person]:
    people: list[Person] = []
    with path.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader, start=2):
            people.append(_row_to_person(row, source_row=i))
    return people


def parse_xlsx(path: Path) -> list[Person]:
    return parse_export_file(path).people


def parse_json(path: Path) -> list[Person]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and "orders" in data:
        return parse_export_file(path).people
    if isinstance(data, dict) and "people" in data:
        items = data["people"]
    elif isinstance(data, list):
        items = data
    else:
        items = [data]
    return [_row_to_person(item, source_row=i + 1) for i, item in enumerate(items)]


def parse_file(path: Path) -> list[Person]:
    return parse_export_file(path).people


def person_from_flags(args: dict[str, str | None]) -> Person:
    return _row_to_person({k: v for k, v in args.items() if v}, source_row=None)


def _shipping_from_row(row: dict[str, Any]) -> str:
    for raw_key, raw_val in row.items():
        key = _norm_key(str(raw_key))
        if key in ("shipping", "shipping address") and raw_val is not None and _clean(raw_val):
            return _clean(raw_val)
    return ""


def extract_shipping_text(path: Path) -> str:
    bundle = parse_export_file(path)
    if bundle.meta.shipping_override:
        return _clean(bundle.meta.shipping_override)
    for batch in bundle.batches:
        if batch.shipping_raw:
            return batch.shipping_raw
    return ""


def parse_shipping_text(text: str) -> ShippingInfo:
    raw = text.strip()
    if not raw:
        return ShippingInfo()

    if _is_local_delivery(raw):
        return ShippingInfo(name="Local Delivery", raw=raw)

    parts = [p.strip() for p in raw.split(",") if p and p.strip()]
    if len(parts) >= 5:
        country = parts[-1] if len(parts) >= 6 else "USA"
        zip_code = parts[-2] if len(parts) >= 6 else parts[-1]
        state = parts[-3] if len(parts) >= 6 else parts[-2]
        city = parts[-4] if len(parts) >= 6 else parts[-3]
        if len(parts) >= 6:
            street = parts[-5]
            name = ", ".join(parts[:-5])
        else:
            name = parts[0]
            street = ", ".join(parts[1:-3])
        return ShippingInfo(
            name=name.strip(),
            street=street.strip(),
            city=city.strip(),
            state=state.strip(),
            zip=zip_code.strip(),
            country=country.strip(),
            raw=raw,
        )

    match = re.search(
        r"^(?P<name>.+?),?\s+(?P<street>\d+.+?),?\s+"
        r"(?P<city>[A-Za-z .'-]+),?\s+(?P<state>[A-Za-z]{2,}|[A-Za-z .'-]+)\s+"
        r"(?P<zip>\d{5}(?:-\d{4})?)",
        raw,
    )
    if match:
        data = match.groupdict()
        return ShippingInfo(
            name=data["name"].strip(),
            street=data["street"].strip(),
            city=data["city"].strip(),
            state=data["state"].strip(),
            zip=data["zip"].strip(),
            raw=raw,
        )

    return ShippingInfo(raw=raw)


def shipping_for_batch(
    batch: OrderBatch,
    bundle: ExportBundle | None,
    cli_override: str = "",
) -> ShippingInfo:
    raw = cli_override or (bundle.meta.shipping_override if bundle else None) or batch.shipping_raw
    return parse_shipping_text(_clean(raw))


@dataclass
class ShippingAddressChoice:
    raw: str
    id_count: int
    order_ids: list[str]
    local_delivery: bool = False


def shipping_choices_from_batches(batches: list[OrderBatch]) -> list[ShippingAddressChoice]:
    groups: dict[str, ShippingAddressChoice] = {}
    for batch in batches:
        if not batch.people:
            continue
        key = batch.shipping_raw.strip() or f"order:{batch.order_id or 'unknown'}"
        if key not in groups:
            groups[key] = ShippingAddressChoice(
                raw=batch.shipping_raw,
                id_count=0,
                order_ids=[],
                local_delivery=batch.local_delivery,
            )
        group = groups[key]
        group.id_count += len(batch.people)
        if batch.order_id and batch.order_id not in group.order_ids:
            group.order_ids.append(batch.order_id)
    return list(groups.values())


def merge_batches(batches: list[OrderBatch], shipping_raw: str = "") -> list[OrderBatch]:
    people: list[Person] = []
    order_ids: list[str] = []
    local_delivery = False
    for batch in batches:
        people.extend(batch.people)
        if batch.order_id:
            order_ids.append(batch.order_id)
        local_delivery = local_delivery or batch.local_delivery

    if not people:
        return batches

    raw = shipping_raw or next((b.shipping_raw for b in batches if b.shipping_raw), "")
    if shipping_raw:
        local_delivery = _is_local_delivery(shipping_raw)

    order_id = order_ids[0] if len(order_ids) == 1 else "merged"
    return [
        OrderBatch(
            order_id=order_id,
            people=people,
            shipping_raw=raw,
            local_delivery=local_delivery,
        )
    ]
